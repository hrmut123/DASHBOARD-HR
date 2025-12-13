import streamlit as st
import pandas as pd
import os
import io
import plotly.express as px
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from streamlit_option_menu import option_menu

# ==========================================
# 1. KONFIGURASI HALAMAN
# ==========================================
st.set_page_config(
    page_title="DASHBOARD HR",
    layout="wide",
    page_icon="🚀",
    initial_sidebar_state="expanded"
)

# ==========================================
# 2. SISTEM LOGIN (SECURITY)
# ==========================================
# 👇 SILAKAN GANTI USERNAME DAN PASSWORD DI BAWAH INI 👇
# Format: "username": "password"
USERS = {
    "kiki": "kiki123",       # Username: admin, Pass: 12345
    "pipin": "pipin123",     # Username: hrd, Pass: 54321
    "adit": "adit123"        # Username: pimpinan, Pass: bos123
}

def check_login(username, password):
    if username in USERS and USERS[username] == password:
        return True
    return False

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

# --- TAMPILAN LOGIN KEREN (FIXED LAYOUT) ---
if not st.session_state['logged_in']:
    st.markdown("""
        <style>
        /* Background Animasi Gradient */
        .stApp {
            background: linear-gradient(-45deg, #0f172a, #1e293b, #0f172a, #2dd4bf);
            background-size: 400% 400%;
            animation: gradient 15s ease infinite;
            color: white;
        }
        @keyframes gradient {
            0% { background-position: 0% 50%; }
            50% { background-position: 100% 50%; }
            100% { background-position: 0% 50%; }
        }

        /* TARGET FORM STREAMLIT LANGSUNG UNTUK EFEK KACA */
        [data-testid="stForm"] {
            background: rgba(255, 255, 255, 0.05);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            padding: 30px;
            border-radius: 20px;
            box-shadow: 0 15px 35px rgba(0, 0, 0, 0.5);
        }

        /* Judul & Logo di dalam Form */
        .login-header {
            text-align: center;
            margin-bottom: 20px;
        }
        .login-icon { font-size: 4rem; margin-bottom: 5px; }
        .login-title {
            font-size: 2.5rem;
            font-weight: 800;
            background: -webkit-linear-gradient(45deg, #38bdf8, #818cf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin: 0;
        }
        .login-subtitle {
            color: #94a3b8;
            font-size: 0.9rem;
            margin-top: 5px;
        }

        /* Input Field */
        div[data-testid="stTextInput"] input {
            background-color: rgba(15, 23, 42, 0.8) !important;
            border: 1px solid #334155 !important;
            color: #f8fafc !important;
            border-radius: 10px !important;
            padding: 10px 15px !important;
        }
        div[data-testid="stTextInput"] input:focus {
            border-color: #38bdf8 !important;
            box-shadow: 0 0 10px rgba(56, 189, 248, 0.3) !important;
        }

        /* Tombol Login */
        div.stButton > button {
            background: linear-gradient(90deg, #3b82f6, #06b6d4);
            color: white;
            font-weight: bold;
            border: none;
            padding: 12px 0;
            border-radius: 50px;
            width: 100%;
            margin-top: 10px;
            transition: all 0.3s;
        }
        div.stButton > button:hover {
            transform: scale(1.02);
            box-shadow: 0 5px 15px rgba(6, 182, 212, 0.4);
        }

        /* Hilangkan elemen pengganggu */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}
        [data-testid="stForm"] span[data-testid="InputInstructions"] { display: none; }
        </style>
    """, unsafe_allow_html=True)

    # Layout Tengah
    col1, col2, col3 = st.columns([1, 1.2, 1])
    
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True) # Spacer atas
        
        # FORM LOGIN (Visual dimasukkan ke dalam form agar menyatu)
        with st.form("login_form"):
            # Header Visual (Icon & Judul)
            st.markdown("""
                <div class="login-header">
                    <div class="login-icon">🔐</div>
                    <div class="login-title">DASHBOARD HR</div>
                    <div class="login-subtitle">Silahkan Login</div>
                </div>
            """, unsafe_allow_html=True)
            
            username = st.text_input("Username", placeholder="ID Pengguna")
            password = st.text_input("Password", type="password", placeholder="Kata Sandi")
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Tombol Submit
            submitted = st.form_submit_button("MASUK SYSTEM")
            
            if submitted:
                if check_login(username, password):
                    st.session_state['logged_in'] = True
                    st.success("Akses Diterima!")
                    st.rerun()
                else:
                    st.error("⛔ Username atau Password Salah")

        # Footer kecil di luar kotak
        st.markdown("""
            <div style='text-align: center; margin-top: 20px; color: #475569; font-size: 12px;'>
                © 2025 HR Management System
            </div>
        """, unsafe_allow_html=True)

    st.stop()

# ==========================================
# 3. CSS DASHBOARD UTAMA (SETELAH LOGIN)
# ==========================================
st.markdown("""
    <style>
    .stApp { background-color: #0f172a; color: #f8fafc; }
    section[data-testid="stSidebar"] { background-color: #1e293b; border-right: 1px solid #334155; }
    
    /* Metrics */
    div[data-testid="metric-container"] { background-color: #1e293b; border: 1px solid #334155; padding: 15px; border-radius: 10px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.3); }
    div[data-testid="metric-container"] label { color: #94a3b8; }
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] { color: #38bdf8; }
    
    /* Tabel */
    th { background-color: #020617 !important; color: #38bdf8 !important; border-bottom: 2px solid #334155 !important; text-align: center !important; }
    td { color: #e2e8f0 !important; background-color: #1e293b !important; text-align: center !important; }
    
    /* Fix Warna Input & Selectbox */
    .stTextInput input, .stNumberInput input, .stDateInput input, .stTextArea textarea { 
        background-color: #334155 !important; color: white !important; border: 1px solid #475569 !important; border-radius: 5px; 
    }
    div[data-testid="stSelectbox"] > div > div {
        background-color: #334155 !important; color: white !important; border: 1px solid #475569 !important; border-radius: 5px;
    }
    div[data-testid="stSelectbox"] div[data-testid="stMarkdownContainer"] p { color: white !important; }

    /* Tombol Dashboard */
    .stButton button { width: 100%; border-radius: 8px; font-weight: bold; border: none; padding: 10px; }
    button[kind="primary"] { background-color: #3b82f6; color: white; }
    button[kind="secondary"] { background-color: #ef4444; color: white; border: 1px solid #dc2626; }
    .streamlit-expanderHeader { background-color: #1e293b !important; color: white !important; border: 1px solid #334155; }
    
    /* Hilangkan instruksi form global */
    [data-testid="stForm"] span[data-testid="InputInstructions"] { display: none; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 4. BACKEND LOGIC
# ==========================================
FILE_EMP = 'data_karyawan.csv'
FILE_ATT = 'data_absensi.csv'
DEFAULT_COLS = ['PT', 'NIK', 'Nama', 'Jabatan', 'Departemen']
ATT_COLS = ['Tanggal', 'NIK', 'Nama', 'Departemen', 'Jenis', 'Keterangan', 'Waktu_Input', 'Durasi']

def clean_dataframe(df):
    df.columns = [str(c).strip() for c in df.columns]
    valid_cols = [c for c in df.columns if 'UNNAMED' not in c.upper() and c != '' and c.lower() != 'nan']
    df = df[valid_cols]
    for col in ['No', 'Ceklist', 'Pilih']:
        if col in df.columns: df = df.drop(columns=[col])
    df = df.fillna("-")
    temp_df = df.replace("-", None)
    df = df[temp_df.notna().any(axis=1)]
    return df

def load_data():
    if os.path.exists(FILE_EMP):
        try:
            df = pd.read_csv(FILE_EMP, dtype=str)
            df = clean_dataframe(df)
            if len(df.columns) == 0: df = pd.DataFrame(columns=DEFAULT_COLS)
        except: df = pd.DataFrame(columns=DEFAULT_COLS)
    else: df = pd.DataFrame(columns=DEFAULT_COLS)

    if os.path.exists(FILE_ATT):
        try:
            df_att = pd.read_csv(FILE_ATT, dtype=str)
            df_att = clean_dataframe(df_att)
            if 'Durasi' not in df_att.columns: df_att['Durasi'] = 1
        except: df_att = pd.DataFrame(columns=ATT_COLS)
    else: df_att = pd.DataFrame(columns=ATT_COLS)
    
    return df, df_att

def save_data(df, df_att):
    df = clean_dataframe(df)
    df_att = clean_dataframe(df_att)
    df.to_csv(FILE_EMP, index=False)
    df_att.to_csv(FILE_ATT, index=False)

def update_original_excel(original_file, df_new, sheet_name, start_row):
    try:
        original_file.seek(0)
        wb = load_workbook(io.BytesIO(original_file.getvalue()))
        if sheet_name not in wb.sheetnames: return None, f"Sheet '{sheet_name}' tidak ditemukan."
        ws = wb[sheet_name]
        clean_df = clean_dataframe(df_new)
        data_rows = clean_df.values.tolist()
        excel_start_row = start_row + 1 
        for i, row_data in enumerate(data_rows):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=i + excel_start_row, column=j + 1)
                cell.value = value
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, "Success"
    except Exception as e: return None, str(e)

def create_colorful_excel(df, title_text):
    output = io.BytesIO()
    clean_df = clean_dataframe(df)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        clean_df.to_excel(writer, index=False, sheet_name='Laporan', startrow=3)
        ws = writer.sheets['Laporan']
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        row_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        font_white = Font(color="FFFFFF", bold=True, size=11)
        font_title = Font(color="1F4E78", bold=True, size=16)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center = Alignment(horizontal="center", vertical="center")
        ws['A1'] = title_text; ws['A1'].font = font_title
        ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        for cell in ws[4]:
            cell.fill = header_fill; cell.font = font_white; cell.alignment = center; cell.border = border
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border; cell.alignment = center
                if cell.row % 2 == 0: cell.fill = row_fill
        for col in ws.columns:
            max_len = 0
            col_let = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                except: pass
            ws.column_dimensions[col_let].width = (max_len + 2) * 1.2
    output.seek(0)
    return output

df_employees, df_attendance = load_data()

# ==========================================
# 5. SIDEBAR NAVIGATION
# ==========================================
with st.sidebar:
    st.markdown("<h1 style='text-align: center; color: #38bdf8;'>⚡ DASHBOARD HR</h1>", unsafe_allow_html=True)
    selected = option_menu(
        menu_title=None,
        options=["Dashboard Karyawan", "Input Absensi", "Laporan Rekap"],
        icons=["people-fill", "clipboard-data", "file-earmark-bar-graph"],
        default_index=0,
        styles={
            "container": {"padding": "0!important", "background-color": "transparent"},
            "icon": {"color": "#38bdf8", "font-size": "18px"}, 
            "nav-link": {"font-size": "15px", "text-align": "left", "margin":"5px", "--hover-color": "#334155", "color": "#e2e8f0"},
            "nav-link-selected": {"background-color": "#3b82f6", "color": "white"},
        }
    )
    st.markdown("---")
    if st.button("🚪 Logout", type="secondary"):
        st.session_state['logged_in'] = False
        st.rerun()
    st.markdown("---")
    st.caption("Mode: Dark Premium")

# ==========================================
# 6. MENU 1: DASHBOARD KARYAWAN
# ==========================================
if selected == "Dashboard Karyawan":
    st.title("📂 Database Karyawan")
    st.markdown("---")
    
    if 'uploaded_template' not in st.session_state: st.session_state['uploaded_template'] = None
    if 'sheet_name_template' not in st.session_state: st.session_state['sheet_name_template'] = ""
    if 'header_row_template' not in st.session_state: st.session_state['header_row_template'] = 6
    if 'show_download' not in st.session_state: st.session_state['show_download'] = False
    
    if 'editor_key_emp' not in st.session_state: st.session_state['editor_key_emp'] = 0
    if 'confirm_del_emp' not in st.session_state: st.session_state['confirm_del_emp'] = False

    if not df_employees.empty:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Pegawai", len(df_employees))
        dept_num = df_employees['Departemen'].nunique() if 'Departemen' in df_employees.columns else 0
        m2.metric("Departemen", dept_num)
        jab_num = df_employees['Jabatan'].nunique() if 'Jabatan' in df_employees.columns else 0
        m3.metric("Jabatan", jab_num)
        m4.metric("Status", "Active")
        st.write("")
        
        has_dept = 'Departemen' in df_employees.columns
        has_jab = 'Jabatan' in df_employees.columns
        if has_dept or has_jab:
            c1, c2 = st.columns(2)
            with c1:
                if has_dept:
                    d_cnt = df_employees['Departemen'].value_counts().head(10).reset_index()
                    d_cnt.columns = ['Departemen', 'Jumlah']
                    fig = px.bar(d_cnt, x='Departemen', y='Jumlah', color='Departemen', title="Top 10 Departemen", template='plotly_dark')
                    fig.update_layout(showlegend=False, height=350, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig, use_container_width=True)
            with c2:
                if has_jab:
                    j_cnt = df_employees['Jabatan'].value_counts().head(10).reset_index()
                    j_cnt.columns = ['Jabatan', 'Jumlah']
                    fig = px.pie(j_cnt, names='Jabatan', values='Jumlah', title="Top 10 Jabatan", hole=0.5, template='plotly_dark')
                    fig.update_layout(height=350, paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig, use_container_width=True)
        st.divider()

    c_up, c_add = st.columns(2)
    with c_up:
        with st.expander("📥 Import Excel (Standard)", expanded=False):
            up_file = st.file_uploader("File .xlsx", type=['xlsx'])
            if up_file:
                st.session_state['uploaded_template'] = up_file
                try:
                    xls = pd.ExcelFile(up_file)
                    idx = 0
                    for i, n in enumerate(xls.sheet_names): 
                        if "DATABASE SESUAI SO".lower() in n.lower(): idx = i; break
                    sh = st.selectbox("Sheet:", xls.sheet_names, index=idx)
                    
                    rw = st.number_input("Header Baris ke:", min_value=1, value=6, help="Baris di Excel dimana nama kolom berada")
                    
                    if st.button("Load Data", type="primary"):
                        df = pd.read_excel(up_file, sheet_name=sh, header=rw-1, dtype=str)
                        df.columns = [str(c).strip().upper() for c in df.columns]
                        rename_map = {
                            "NO. INDUK": "NIK", "NIK/NRP": "NIK", "NO INDUK": "NIK",
                            "NAMA LENGKAP": "Nama", "NAMA KARYAWAN": "Nama", "NAMA": "Nama",
                            "JABATAN": "Jabatan", "POSISI": "Jabatan", "JABATAN BARU": "Jabatan",
                            "DEPARTEMEN": "Departemen", "DIVISI": "Departemen", "BAGIAN": "Departemen", "DEPARTEMEN BARU": "Departemen", "DEPT": "Departemen",
                            "PT": "PT", "PERUSAHAAN": "PT", "ENTITY": "PT", "PT BARU": "PT"
                        }
                        df.rename(columns=rename_map, inplace=True)
                        df = clean_dataframe(df)
                        df_employees = df; save_data(df_employees, df_attendance)
                        st.session_state['sheet_name_template'] = sh; st.session_state['header_row_template'] = rw
                        st.session_state['show_download'] = False
                        st.session_state['editor_key_emp'] += 1 
                        st.success(f"Loaded {len(df)} rows."); st.rerun()
                except Exception as e: st.error(f"Error: {e}")

    with c_add:
        with st.expander("➕ Tambah Manual"):
            cols = [c for c in df_employees.columns if c not in ['No','Ceklist','Pilih']] or DEFAULT_COLS
            with st.form("add"):
                v = {}
                cg = st.columns(2)
                for i, col in enumerate(cols):
                    with cg[i%2]: v[col] = st.text_input(col)
                if st.form_submit_button("Simpan"):
                    if any(v.values()):
                        df_employees = pd.concat([df_employees, pd.DataFrame([v])], ignore_index=True)
                        save_data(df_employees, df_attendance)
                        st.session_state['editor_key_emp'] += 1 
                        st.success("OK"); st.rerun()

    if not df_employees.empty:
        st.subheader("📝 Editor Data & Hapus")
        search = st.text_input("🔍 Filter Nama/NIK:")
        if search:
            df_show = df_employees[df_employees.astype(str).apply(lambda x: x.str.contains(search, case=False)).any(axis=1)].copy()
        else:
            df_show = df_employees.copy()
        
        df_show.insert(0, "Pilih", False)
        
        edited = st.data_editor(
            df_show, 
            num_rows="dynamic", 
            use_container_width=True, 
            hide_index=True, 
            height=450,
            column_config={"Pilih": st.column_config.CheckboxColumn("Hapus?", width="small")},
            key=f"editor_emp_{st.session_state['editor_key_emp']}"
        )
        
        b1, b2, b3 = st.columns([1, 1, 3])
        
        with b1:
            if st.button("💾 Simpan Perubahan", type="primary"):
                if search: st.warning("Harap hapus kata kunci pencarian.")
                else: 
                    final_df = edited.drop(columns=['Pilih'])
                    save_data(final_df, df_attendance)
                    st.success("Tersimpan!")
                    st.session_state['show_download'] = True
                    st.rerun()
        
        with b2:
            if st.button("🗑️ Hapus Terpilih", type="secondary"):
                if search:
                    st.warning("Hapus filter dulu.")
                else:
                    checked = edited[edited['Pilih'] == True]
                    if checked.empty: st.info("Pilih data dulu.")
                    else: st.session_state['confirm_del_emp'] = True

        with b3:
            if st.session_state.get('show_download', False):
                if st.session_state['uploaded_template']:
                    try:
                        out_buffer, status = update_original_excel(
                            st.session_state['uploaded_template'], 
                            df_employees, 
                            st.session_state['sheet_name_template'], 
                            st.session_state['header_row_template']
                        )
                        if out_buffer:
                            st.download_button(label="📥 Download File Update (Excel Asli)", data=out_buffer, file_name="SO_MUT_UPDATED.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
                        else: st.error(status)
                    except Exception as e: st.error(f"Error: {e}")
                else:
                    xl = create_colorful_excel(df_employees, "DATABASE KARYAWAN")
                    st.download_button("📥 Download Excel Baru", xl, "Database.xlsx")

        if st.session_state['confirm_del_emp']:
            rows_to_delete = edited[edited['Pilih'] == True]
            st.error(f"⚠️ HAPUS {len(rows_to_delete)} DATA PERMANEN?")
            cy, cn = st.columns([1, 1])
            with cy:
                if st.button("✅ Ya, Hapus"):
                    rows_to_keep = edited[edited['Pilih'] == False]
                    final_df = rows_to_keep.drop(columns=['Pilih'])
                    save_data(final_df, df_attendance)
                    st.session_state['confirm_del_emp'] = False
                    st.session_state['editor_key_emp'] += 1 
                    st.success("Terhapus!"); st.rerun()
            with cn:
                if st.button("❌ Batal"):
                    st.session_state['confirm_del_emp'] = False
                    st.session_state['editor_key_emp'] += 1 
                    st.rerun()

    else: st.info("Database kosong.")

# ==========================================
# 6. INPUT ABSENSI
# ==========================================
elif selected == "Input Absensi":
    st.title("📝 Presensi Harian")
    st.markdown("---")
    
    if 'editor_key_att' not in st.session_state: st.session_state['editor_key_att'] = 0
    if 'confirm_del_att' not in st.session_state: st.session_state['confirm_del_att'] = False

    if df_employees.empty: st.error("Database Kosong.")
    else:
        c1, c2 = st.columns([1, 2])
        with c1:
            with st.container(border=True):
                st.info("Form Input")
                cols = df_employees.columns
                cnik = next((c for c in cols if 'NIK' in c.upper()), cols[0])
                cnm = next((c for c in cols if 'NAMA' in c.upper()), cols[1])
                mst = df_employees[[cnik, cnm]].drop_duplicates().dropna()
                opts = [f"{r[cnik]} - {r[cnm]}" for _, r in mst.iterrows()]
                
                sel = st.selectbox("Karyawan:", opts)
                
                # --- UPDATE NAMA IZIN ---
                opsi_absen = ["Sakit (Ada Surat)", "Sakit (Tanpa Surat)", "Izin (Resmi)", "Izin (Tidak Resmi)", "Cuti", "Alpha"]
                jenis = st.selectbox("Keterangan Absen:", opsi_absen)
                
                st.markdown("**Rentang Waktu:**")
                col_tgl1, col_tgl2 = st.columns(2)
                with col_tgl1: tgl_mulai = st.date_input("Dari Tanggal:", datetime.now())
                with col_tgl2: tgl_selesai = st.date_input("Sampai Tanggal:", datetime.now())
                ket = st.text_area("Detail Keterangan:", height=80)
                
                if st.button("Simpan", type="primary", use_container_width=True):
                    if tgl_selesai < tgl_mulai: st.error("❌ Tanggal Sampai tidak boleh lebih kecil.")
                    else:
                        nik_val = sel.split(" - ")[0]
                        nm_val = sel.split(" - ")[1]
                        dpt_val = "-"
                        cdep = next((c for c in cols if 'DEP' in c.upper()), None)
                        if cdep:
                            tmp = df_employees[df_employees[cnik] == nik_val]
                            if not tmp.empty: dpt_val = tmp.iloc[0][cdep]
                        
                        durasi_hari = (tgl_selesai - tgl_mulai).days + 1
                        
                        if durasi_hari > 1:
                            tgl_str = f"{tgl_mulai.strftime('%Y-%m-%d')} s/d {tgl_selesai.strftime('%Y-%m-%d')}"
                        else:
                            tgl_str = tgl_mulai.strftime('%Y-%m-%d')

                        new_row = {
                            'Tanggal': tgl_str, 'NIK': nik_val, 'Nama': nm_val, 'Departemen': dpt_val, 
                            'Jenis': jenis, 'Keterangan': ket, 'Durasi': durasi_hari,
                            'Waktu_Input': datetime.now().strftime("%Y-%m-%d %H:%M")
                        }
                        
                        df_attendance = pd.concat([df_attendance, pd.DataFrame([new_row])], ignore_index=True)
                        save_data(df_employees, df_attendance)
                        st.session_state['editor_key_att'] += 1 
                        st.success(f"Berhasil! (Durasi: {durasi_hari} hari)"); st.rerun()

        with c2:
            st.subheader("Riwayat & Edit")
            if not df_attendance.empty:
                df_show = df_attendance.sort_values('Waktu_Input', ascending=False).reset_index(drop=True)
                df_show.insert(0, "Pilih", False)
                
                edited_history = st.data_editor(
                    df_show,
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "Pilih": st.column_config.CheckboxColumn("Hapus?", width="small"),
                        "Waktu_Input": st.column_config.TextColumn("Waktu", disabled=True),
                        "Nama": st.column_config.TextColumn("Nama", disabled=True),
                        "Jenis": st.column_config.TextColumn("Jenis", disabled=True),
                        "Durasi": st.column_config.NumberColumn("Hari", disabled=True, width="small")
                    },
                    key=f"editor_att_{st.session_state['editor_key_att']}"
                )
                
                col_del, col_space = st.columns([1, 3])
                with col_del:
                    if st.button("🗑️ Hapus Data Terpilih", type="secondary"):
                        checked = edited_history[edited_history['Pilih'] == True]
                        if checked.empty: st.info("Pilih data dulu!")
                        else: st.session_state['confirm_del_att'] = True

                if st.session_state['confirm_del_att']:
                    checked = edited_history[edited_history['Pilih'] == True]
                    st.error(f"⚠️ HAPUS {len(checked)} DATA ABSENSI?")
                    cy, cn = st.columns([1, 1])
                    with cy:
                        if st.button("✅ Ya, Hapus", key="del_att_yes"):
                            rows_to_keep = edited_history[edited_history['Pilih'] == False]
                            rows_to_keep = rows_to_keep.drop(columns=['Pilih'])
                            df_attendance = rows_to_keep
                            save_data(df_employees, df_attendance)
                            st.session_state['confirm_del_att'] = False
                            st.session_state['editor_key_att'] += 1 
                            st.success("Terhapus!"); st.rerun()
                    with cn:
                        if st.button("❌ Batal", key="del_att_no"):
                            st.session_state['confirm_del_att'] = False
                            st.session_state['editor_key_att'] += 1 
                            st.rerun()
            else: 
                st.info("Belum ada data absensi.")

# ==========================================
# 8. REKAP LAPORAN
# ==========================================
elif selected == "Laporan Rekap":
    st.title("📊 Laporan Bulanan")
    st.markdown("---")
    if df_employees.empty: st.warning("Database Kosong.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        bln = c1.selectbox("Bulan", range(1,13), index=datetime.now().month-1)
        thn = c2.number_input("Tahun", value=datetime.now().year)
        hk = c3.number_input("Hari Kerja", value=26)
        
        df_att_show = df_attendance.copy()
        if not df_att_show.empty:
            df_att_show['Tgl_Filter'] = df_att_show['Tanggal'].astype(str).str.slice(0, 10)
            df_att_show['Tgl_Filter'] = pd.to_datetime(df_att_show['Tgl_Filter'], errors='coerce')
            
            mask = (df_att_show['Tgl_Filter'].dt.month == bln) & (df_att_show['Tgl_Filter'].dt.year == thn)
            fil = df_att_show[mask]
            
            fil['Durasi'] = pd.to_numeric(fil['Durasi'], errors='coerce').fillna(1)
            rekap_jenis = fil.groupby('Jenis')['Durasi'].sum()
            
            m_cols = st.columns(6)
            
            # --- UPDATE KATEGORI AGAR SESUAI DENGAN INPUT ---
            categories = ["Sakit (Ada Surat)", "Sakit (Tanpa Surat)", "Izin (Resmi)", "Izin (Tidak Resmi)", "Cuti", "Alpha"]
            
            for i, cat in enumerate(categories):
                val = rekap_jenis.get(cat, 0)
                m_cols[i].metric(cat, int(val))
            
            st.divider()

            pivot_absen = fil.pivot_table(index='NIK', columns='Jenis', values='Durasi', aggfunc='sum', fill_value=0).reset_index()
            cols = df_employees.columns
            cnik = next((c for c in cols if 'NIK' in c.upper()), cols[0])
            cnm = next((c for c in cols if 'NAMA' in c.upper()), cols[1])
            cdep = next((c for c in cols if 'DEP' in c.upper()), None)
            cpt = next((c for c in cols if 'PT' in c.upper()), None)
            
            cols_to_fetch = [cnik, cnm]
            if cdep: cols_to_fetch.append(cdep)
            if cpt: cols_to_fetch.append(cpt)
            
            mst = df_employees[cols_to_fetch].drop_duplicates()
            
            rename_map = {cnik: 'NIK', cnm: 'Nama'}
            if cdep: rename_map[cdep] = 'Departemen'
            if cpt: rename_map[cpt] = 'PT'
            mst = mst.rename(columns=rename_map)
            
            if 'Departemen' not in mst.columns: mst['Departemen'] = "-"
            if 'PT' not in mst.columns: mst['PT'] = "-"
            
            mst['NIK'] = mst['NIK'].astype(str); pivot_absen['NIK'] = pivot_absen['NIK'].astype(str)
            fin = pd.merge(mst, pivot_absen, on='NIK', how='left')
            
            for cat in categories:
                if cat not in fin.columns: fin[cat] = 0
            fin = fin.fillna(0)
            
            fin['Total_Absen'] = fin[categories].sum(axis=1)
            fin['Total_Hadir'] = (hk - fin['Total_Absen']).clip(lower=0)
            fin['Persentase'] = ((fin['Total_Hadir']/hk)*100).round(1).astype(str) + '%'
            
            final_cols = ['PT', 'NIK', 'Nama', 'Departemen'] + categories + ['Total_Absen', 'Total_Hadir', 'Persentase']
            fin = fin[final_cols]
            
            st.dataframe(fin, use_container_width=True, hide_index=True)
            
            xl = create_colorful_excel(fin, f"REKAP {bln}/{thn}")
            st.download_button("📥 Download Laporan (Excel)", xl, f"Rekap_{bln}_{thn}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
            
            st.divider()
            
            if not fil.empty:
                g1, g2 = st.columns(2)
                with g1:
                    df_pie = rekap_jenis.reset_index()
                    df_pie.columns = ['Jenis', 'Total Hari']
                    fig = px.pie(df_pie, names='Jenis', values='Total Hari', title="Proporsi Ketidakhadiran (Hari)", hole=0.4, template="plotly_dark")
                    fig.update_layout(paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig, use_container_width=True)
                with g2:
                    top = fin[fin['Total_Absen']>0].sort_values('Total_Absen', ascending=False).head(5)
                    fig = px.bar(top, x='Total_Absen', y='Nama', orientation='h', title="Top 5 Paling Sering Absen", template="plotly_dark")
                    fig.update_layout(paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig, use_container_width=True)
        else: st.info("Belum ada data absensi.")